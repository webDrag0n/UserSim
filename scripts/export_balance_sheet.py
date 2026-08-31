"""从 usersim/world/catalog.py 导出游戏数值平衡风格 Excel 配表。

用法：.venv/bin/python scripts/export_balance_sheet.py
输出：balance-sheet/UserSim数值配表.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from usersim.config import load_system_config
from usersim.world.catalog import (
    DISTURBANCES, ECONOMY, MEAL_TIERS, PROFESSIONS, RECOVERY_ACTIONS,
    SLEEP_TIERS, TEMPLATE_EVENTS, all_variants, get_venues,
)

DIM_LABELS = {"valence": "心情", "energy": "精力", "satiety": "饱腹", "stress": "压力"}

# ---------- 样式 ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="111827")
NOTE_FONT = Font(name="微软雅黑", size=9, color="6B7280")
TIER_FILLS = {
    "平价": PatternFill("solid", fgColor="ECFDF5"),
    "中档": PatternFill("solid", fgColor="FFFBEB"),
    "高档": PatternFill("solid", fgColor="FEF2F2"),
    "应付": PatternFill("solid", fgColor="F3F4F6"),
    "日常": PatternFill("solid", fgColor="ECFDF5"),
    "品质": PatternFill("solid", fgColor="FFFBEB"),
    "劣质": PatternFill("solid", fgColor="FEF2F2"),
}
ALT_FILL = PatternFill("solid", fgColor="F9FAFB")
THIN = Border(*[Side(style="thin", color="D1D5DB")] * 4)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fmt_effect(eff: dict) -> str:
    """效果 dict → 可读字符串：心情+0.06 压力-0.04；pull 类 → 饱腹→0.70(×0.75)"""
    parts = []
    for k, v in eff.items():
        label = DIM_LABELS.get(k, k)
        if isinstance(v, dict) and "pull" in v:
            parts.append(f"{label}→{v['pull'][0]}(×{v['pull'][1]})")
        else:
            parts.append(f"{label}{'+' if v > 0 else ''}{v:g}")
    return " ".join(parts) if parts else "—"


def style_sheet(ws, headers, widths, title, note, rows, tier_col=None, center_cols=()):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c = ws.cell(2, 1, note)
    c.font = NOTE_FONT
    c.alignment = LEFT
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    for j, h in enumerate(headers, 1):
        cell = ws.cell(3, j, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    for i, row in enumerate(rows):
        r = i + 4
        for j, val in enumerate(row, 1):
            cell = ws.cell(r, j, val)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = CENTER if j in center_cols else LEFT
            if tier_col and j == tier_col and val in TIER_FILLS:
                cell.fill = TIER_FILLS[val]
            elif i % 2 == 1:
                cell.fill = ALT_FILL
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _build_anthro_sheets(wb) -> None:
    """拟人化参数三表（docs/11）。仅在 sheet 缺失时创建，保护用户编辑。"""
    from usersim.world.anthro import HABITUATION_DEFAULTS

    if "习惯化曲线" not in wb.sheetnames:
        ws = wb.create_sheet("习惯化曲线")
        intent = {
            "吃好吃的": "好吃的耐受较强，一两天恢复", "好好休息": "身体适应快",
            "出门走走": "散步容易腻，半天即恢复", "短途旅行": "旅行记忆持久，半个月才再想去",
            "运动健身": "习惯化浅", "宅家回血": "纯宅很快就无聊",
            "文化看展": "展览很快腻，但隔几天又很想看", "咖啡小憩": "小额慰藉天天喝也还行",
            "音乐放松": "情绪价值类", "朋友小聚": "社交恢复前快后慢",
            "自然放空": "免费强降压但容易腻", "自定义活动": "兜底",
        }
        style_sheet(ws, ["动作", "w_min(连续重复下限)", "τ(恢复常数·时段)", "曲线类型", "设计意图"],
                    [16, 18, 18, 10, 40],
                    "习惯化曲线（边际效益）：w(Δt)=1-(1-w_min)·c(Δt)，Δt=距上次执行时段数",
                    "exp=指数恢复；sqrt=前快后慢（社交类）；s=S型（爱好）。无计时器，按上次执行时间差查表。",
                    [[name, w, t, c, intent.get(name, "")] for name, (w, t, c) in HABITUATION_DEFAULTS.items()],
                    center_cols=(2, 3, 4))

    if "需求参数" not in wb.sheetnames:
        ws = wb.create_sheet("需求参数")
        style_sheet(ws, ["需求", "累积规则", "满足事件", "驱动力曲线 u(x)", "满足曲线 s(x)", "设计意图"],
                    [10, 30, 22, 24, 26, 36],
                    "需求动力学（认知动力学曲线，非全部单调）",
                    "u(x) 驱动求助倾向；s(x) 调制满足事件的效果权重。刺激需求为倒 U（太少无聊/太多过载）。",
                    [
                        ["饥饿", "由饱腹推导（低饱腹加速）", "吃好吃的/三餐", "u=((1-x)/0.6)^1.5", "s=1+1.5u", "越饿吃得越香"],
                        ["社交", "+0.01/时段（外向×1.6）", "朋友小聚/应酬", "u=x²", "s=1+0.8u", "社交电池"],
                        ["刺激", "单调-0.01/时段，新异+0.12", "看展/旅行/新地方", "倒U: u=1-(2x-1)²", "s=0.6+0.8·(1-|2x-1|)", "过载也烦躁"],
                        ["成就", "备考/截止逼近陡增", "刷题/工作推进", "u=x^2.5", "完成时释放∝x", "deadline 效应"],
                    ], center_cols=(1,))

    if "人格调节" not in wb.sheetnames:
        ws = wb.create_sheet("人格调节")
        style_sheet(ws, ["维度", "规则", "设计意图"],
                    [12, 70, 36],
                    "人格调节（大五生效）",
                    "人格不再只是冻结摆设：同一事件对不同人效果不同。",
                    [
                        ["外向性", "社交事件精力×(1+1.2E)/(1.6-1.2E)；E>0.7 额外心情+0.03", "社交电池：内向耗电、外向回血"],
                        ["神经质", "压力事件效果×(1+N-0.5)；压力均值回归速率×(1-0.4N)", "高神经质更敏感、恢复更慢"],
                        ["开放性", "文化/新异事件效果×(0.7+0.6O)", "高开放性更享受新刺激"],
                    ], center_cols=(1,))


def _init_missing_sheets(path: Path) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    before = set(wb.sheetnames)
    _build_anthro_sheets(wb)
    if set(wb.sheetnames) != before:
        wb.save(path)
        print(f"已补充缺失 sheet: {sorted(set(wb.sheetnames) - before)}")


def main() -> None:
    out_dir = Path(__file__).resolve().parent.parent / "balance-sheet"
    out_dir.mkdir(exist_ok=True)
    out = out_dir / "UserSim数值配表.xlsx"
    if out.exists():
        # 文件已存在：只补缺失 sheet（拟人化三表等），不覆盖任何用户编辑
        _init_missing_sheets(out)
        print(f"已保留现有内容: {out}")
        return
    wb = Workbook()
    cfg = load_system_config()

    # ---------- Sheet 1 恢复事件表（事件元信息：无地点无效果） ----------
    ws = wb.active
    ws.title = "恢复事件表"
    headers = ["事件ID", "事件", "类别", "默认时长(时段)", "设计意图"]
    rows = [[a["id"], a["action"], a["category"], a.get("default_span", 1), a["design_intent"]]
            for a in RECOVERY_ACTIONS]
    style_sheet(ws, headers, [8, 10, 7, 12, 60],
                "恢复事件表（事件只携带元信息；价格与效果在「地点支持表」）",
                "事件不再携带地点维度与效果：一次「在某地点做某事件」的价格/时长/效果由地点支持记录自带（venues.json 的 supports）。世界只信配表，LLM 只能选事件与地点。",
                rows, center_cols=(1, 3, 4))

    # ---------- Sheet 1b 地点支持表（venues × supports flatten） ----------
    ws1b = wb.create_sheet("地点支持表")
    headers1b = ["变体ID", "事件", "地点", "类目", "菜系", "价格¥", "时长(时段)",
                 "效果", "替代模板餐", "地点设计意图"]
    venue_by_id = {v["id"]: v for v in get_venues()}
    rows1b = []
    for a, v in all_variants():
        if a.get("id") in ("MEAL", "SLEEP"):
            continue  # 日常升级档在「日常事件配表」
        vn = venue_by_id.get(v.get("id"), {})
        rows1b.append([v["vid"], a["action"], v["location"], v.get("category", ""),
                       v.get("cuisine", ""), v["cost"], v.get("span", 1),
                       fmt_effect(v.get("effect", {})),
                       "是" if v.get("replaces_meal") else "",
                       vn.get("design_intent", "")])
    style_sheet(ws1b, headers1b, [13, 10, 24, 7, 8, 8, 10, 26, 10, 40],
                "地点支持表（vid = 事件id@地点id；同一地点可支持多事件多条目）",
                "效果为事件总值（旧档位为 base+weight 合计，逐字迁移），多时段按 span 摊销（pull 类除外）；价格一次性消耗。餐饮场所替代当日模板餐（replaces_meal）。",
                rows1b, center_cols=(1, 2, 4, 5, 6, 7, 9))

    # ---------- Sheet 2 日常事件（进餐/睡眠） ----------
    ws2 = wb.create_sheet("日常事件配表")
    headers2 = ["档位ID", "名称", "类别", "档位", "价格¥", "效果", "设计意图"]
    rows2 = (
        [[m["vid"], m["name"], "进餐", m["tier"], m["cost"], fmt_effect(m["effect"]), m["design_intent"]] for m in MEAL_TIERS]
        + [[s["vid"], s["name"], "睡眠", s["tier"], s["cost"], fmt_effect(s["effect"]), s["design_intent"]] for s in SLEEP_TIERS]
    )
    style_sheet(ws2, headers2, [8, 22, 7, 7, 8, 26, 44],
                "日常事件配表（进餐 / 睡眠）",
                "进餐与睡眠不再是全局参数，而是事件：日常档进模板日程（M1 家常 ¥10/顿、S1 正常睡眠免费），可被助手/用户升级为品质档。pull 效果 = 拉向准稳态（目标×速率），不按时长摊销。",
                rows2, tier_col=4, center_cols=(1, 3, 4, 5))

    # ---------- Sheet 2b 自定义活动类目 ----------
    from usersim.world.catalog import CUSTOM_ACTIVITIES
    ws2b = wb.create_sheet("自定义活动类目")
    headers2b = ["类目ID", "规范名", "价格¥", "关键词（命中即归一化）", "效果", "设计意图"]
    rows2b = [
        [c["id"], c["name"], c["cost"], "、".join(c["keywords"]) or "（兜底：未命中任何关键词）",
         fmt_effect(c["effect"]), c["design_intent"]]
        for c in CUSTOM_ACTIVITIES
    ]
    style_sheet(ws2b, headers2b, [8, 12, 8, 44, 22, 32],
                "自定义活动类目（LLM 目录外活动的归一化表）",
                "LLM 发明的自由措辞按关键词归入规范类目：名称/效果/价格由世界裁定，同类活动在日程图与日志中合并为一行；原称保留在目标字段备查。",
                rows2b, center_cols=(1, 3))

    # ---------- Sheet 3 职业收入表 ----------
    ws3 = wb.create_sheet("职业收入表")
    headers3 = ["职业", "收入¥/工作时段", "日收入¥(×2)", "说明"]
    rows3 = [[p["archetype"], p["income_per_slot"], p["income_per_slot"] * 2, p["note"]] for p in PROFESSIONS]
    style_sheet(ws3, headers3, [20, 14, 12, 44],
                "职业收入表（收入按职业分档）",
                "收入在每个工作时段（上午/下午）结算；加班事件另有 ¥150。低收入职业必须依赖免费恢复档，经济压力本身成为难度维度。",
                rows3, center_cols=(2, 3))

    # ---------- Sheet 4 扰动事件配表 ----------
    ws4 = wb.create_sheet("扰动事件配表")
    headers4 = ["事件ID", "名称", "地点", "价格¥", "收入¥", "效果", "设计意图"]
    rows4 = [
        [d["id"], d["name"], d["location"], d["cost"], d["income"], fmt_effect(d["effect"]), d["design_intent"]]
        for d in DISTURBANCES
    ]
    style_sheet(ws4, headers4, [8, 14, 10, 8, 8, 26, 48],
                "扰动事件配表（泊松流：每天 62% 概率一次）",
                "扰动是系统的控制扰动输入 d(t)；加班有收入（钱换命），邀约为难得的正向扰动但也有代价。",
                rows4, center_cols=(1, 4, 5))

    # ---------- Sheet 5 模板与动力学 ----------
    ws5 = wb.create_sheet("模板与动力学")
    headers5 = ["ID", "名称", "时段", "地点", "隐含效果（自然动力学结算）"]
    rows5 = [[t["id"], t["name"], t["slot"], t["location"], t["implicit_effect"]] for t in TEMPLATE_EVENTS]
    style_sheet(ws5, headers5, [6, 10, 14, 10, 60],
                "模板事件表（作息铺底：工作/休整）",
                "工作/休整的作用融入自然动力学，避免双重计数；进餐与睡眠已移入「日常事件配表」。",
                rows5, center_cols=(1,))

    # ---------- Sheet 5 系列事件配表 ----------
    from usersim.world.series import SERIES_TYPES
    ws5 = wb.create_sheet("系列事件配表")
    headers5 = ["系列", "项目", "名称", "价格¥", "收入¥", "时长(时段)", "效果", "备注"]
    rows5 = []
    part_names = {"sleep": "宿", "meals": "餐", "daily_pool": "子事件", "evening_pool": "晚间", "transit": "交通", "aftereffect": "后效", "final_event": "终章"}
    for stype, sdef in SERIES_TYPES.items():
        meta = f"{sdef['icon']}{sdef['name']}（{sdef['duration_range'][0]}~{sdef['duration_range'][1]}天 · {'规划' if sdef['source']=='planned' else '强制'} · {'停收入' if sdef['suppress_income'] else '有收入'}）"
        def srow(part, it, span=1):
            rows5.append([meta if not rows5 or rows5[-1][0] != meta else "", part, it["name"],
                          it.get("cost", 0), it.get("income", 0), span, fmt_effect(it.get("effect", {})), it.get("note", "")])
        srow(part_names["sleep"], sdef["sleep"])
        srow(part_names["meals"], sdef["meal"])
        for p in sdef["daily_pool"]:
            srow(part_names["daily_pool"], p, p.get("span", 1))
        for p in sdef.get("evening_pool", []):
            srow(part_names["evening_pool"], p)
        if sdef.get("transit"):
            srow(part_names["transit"], sdef["transit"])
        if sdef.get("final_event"):
            srow(part_names["final_event"], sdef["final_event"])
        if sdef.get("aftereffect"):
            rows5.append([ "", part_names["aftereffect"], sdef["aftereffect"]["name"], 0, 0,
                          f"{sdef['aftereffect']['days']}天", fmt_effect(sdef["aftereffect"]["effect"]), "系列结束后的持续影响"])
    style_sheet(ws5, headers5, [34, 7, 18, 8, 8, 10, 30, 30],
                "系列事件配表（跨多天剧情块：行程单一次物化）",
                "系列区间内日常模板被覆盖：旅行停工作停收入、异地餐宿效果不同于日常；出差有补贴但认床应酬多；宅家休假有空虚机制；备考冲刺强制替换休闲。",
                rows5, center_cols=(2, 4, 5, 6))

    # ---------- Sheet 6 经济与全局参数 ----------
    ws6 = wb.create_sheet("经济与全局参数")
    dyn = cfg.dynamics.to_dict()
    st = cfg.state.to_dict()
    rows6 = [
        ["初始金钱", ECONOMY["initial_money"], "用户初始金钱"],
        ["加班收入", ECONOMY["overtime_income"], "临时加班事件的额外收入（工资按职业表）"],
        ["负债压力", ECONOMY["debt_stress_per_slot"], "金钱为负时每时段压力惩罚"],
        ["目标·心情 valence", st["targets"]["valence"], "内心平和带设定点（单侧误差）"],
        ["目标·精力 energy", st["targets"]["energy"], ""],
        ["目标·饱腹 satiety", st["targets"]["satiety"], ""],
        ["目标·压力 stress", st["targets"]["stress"], "压力为反向指标（高于目标才算误差）"],
        ["平和带半宽 band", st["band"], "单侧误差容忍度"],
        ["饱腹消耗/时段", dyn["satiety_drain_per_slot"], "新陈代谢"],
        ["工作压力增速/时段", dyn["work_stress_per_slot"], "工作日工作时段"],
        ["工作精力消耗/时段", dyn["work_energy_drain"], ""],
        ["休息降压/时段", dyn["rest_stress_relief"], "晚上"],
        ["反弹阈值", dyn["rebound_threshold"], "压力低于此值触发积压反弹"],
        ["反弹倍率", dyn["rebound_multiplier"], "反弹时工作效果倍率"],
        ["心情耦合速率", dyn["valence_coupling_rate"], "心情向 v_eq 漂移"],
    ]
    style_sheet(ws6, ["参数", "值", "说明"], [22, 12, 60], "经济与全局参数",
                "与 config/system.toml 对应；进餐/睡眠参数已移入「日常事件配表」。评估为单侧误差（健康维低于目标/压力高于目标才算偏差）。",
                rows6, center_cols=(2,))

    _build_anthro_sheets(wb)
    wb.save(out)
    print(f"已生成: {out}")


if __name__ == "__main__":
    main()

"""从 Excel 配表迁移到 JSON 配置文件。

用法：.venv/bin/python scripts/migrate_excel_to_json.py
输入：balance-sheet/UserSim数值配表.xlsx
输出：config/balance/*.json (12 个文件)
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = PROJECT_ROOT / "balance-sheet" / "UserSim数值配表.xlsx"
CONFIG_DIR = PROJECT_ROOT / "config" / "balance"

_DIM_EN = {"心情": "valence", "精力": "energy", "饱腹": "satiety", "压力": "stress"}


def parse_effect(text: str) -> dict:
    """'心情+0.06 压力-0.04' 或 '饱腹→0.70(×0.75)' → effect dict."""
    import re
    out: dict = {}
    if not text or text == "—":
        return out
    pull_re = re.compile(r"^(心情|精力|饱腹|压力)→([\d.]+)\(×([\d.]+)\)$")
    eff_re = re.compile(r"^(心情|精力|饱腹|压力)([+-][\d.]+)$")

    for part in str(text).split():
        m = pull_re.match(part)
        if m:
            out[_DIM_EN[m.group(1)]] = {"pull": [float(m.group(2)), float(m.group(3))]}
            continue
        m = eff_re.match(part)
        if m:
            out[_DIM_EN[m.group(1)]] = float(m.group(2))
    return out


def rows(ws, min_row=4):
    """Extract rows from worksheet starting at min_row."""
    return [list(r) for r in ws.iter_rows(min_row=min_row, values_only=True)
            if r and r[0] is not None]


def migrate_recovery_actions(wb) -> list[dict]:
    """恢复事件配表 → recovery_actions.json"""
    if "恢复事件配表" not in wb.sheetnames:
        return []

    actions: dict[str, dict] = {}
    for row in rows(wb["恢复事件配表"]):
        aid, action = str(row[0]), str(row[1])
        a = actions.setdefault(aid, {
            "id": aid,
            "action": action,
            "category": str(row[2]),
            "base_effect": parse_effect(row[8]),
            "design_intent": str(row[11] or ""),
            "variants": []
        })
        a["variants"].append({
            "vid": str(row[3]),
            "location": str(row[4]),
            "tier": str(row[5]),
            "cost": float(row[6]),
            "span": int(row[7]),
            "weight": parse_effect(row[9]),
        })
    return list(actions.values())


def migrate_meal_tiers(wb) -> list[dict]:
    """日常事件配表 → meal_tiers.json"""
    if "日常事件配表" not in wb.sheetnames:
        return []

    meals = []
    for row in rows(wb["日常事件配表"]):
        category = str(row[2])
        if category == "进餐":
            meals.append({
                "vid": str(row[0]),
                "name": str(row[1]),
                "tier": str(row[3]),
                "cost": float(row[4]),
                "effect": parse_effect(row[5]),
                "design_intent": str(row[6] or "")
            })
    return meals


def migrate_sleep_tiers(wb) -> list[dict]:
    """日常事件配表 → sleep_tiers.json"""
    if "日常事件配表" not in wb.sheetnames:
        return []

    sleeps = []
    for row in rows(wb["日常事件配表"]):
        category = str(row[2])
        if category == "睡眠":
            sleeps.append({
                "vid": str(row[0]),
                "name": str(row[1]),
                "tier": str(row[3]),
                "cost": float(row[4]),
                "effect": parse_effect(row[5]),
                "design_intent": str(row[6] or "")
            })
    return sleeps


def migrate_custom_activities(wb) -> list[dict]:
    """自定义活动类目 → custom_activities.json"""
    if "自定义活动类目" not in wb.sheetnames:
        return []

    activities = []
    for row in rows(wb["自定义活动类目"]):
        keywords_str = str(row[3] or "")
        keywords = [k.strip() for k in keywords_str.split("、") if k.strip()]
        if "（兜底" in keywords_str:
            keywords = []

        activities.append({
            "id": str(row[0]),
            "name": str(row[1]),
            "cost": float(row[2]),
            "keywords": keywords,
            "effect": parse_effect(row[4]),
            "design_intent": str(row[5] or "")
        })
    return activities


def migrate_professions(wb) -> list[dict]:
    """职业收入表 → professions.json"""
    if "职业收入表" not in wb.sheetnames:
        return []

    profs = []
    for row in rows(wb["职业收入表"]):
        profs.append({
            "archetype": str(row[0]),
            "income_per_slot": int(row[1]),
            "note": str(row[3] or "")
        })
    return profs


def migrate_disturbances(wb) -> list[dict]:
    """扰动事件配表 → disturbances.json"""
    if "扰动事件配表" not in wb.sheetnames:
        return []

    dists = []
    for row in rows(wb["扰动事件配表"]):
        dists.append({
            "id": str(row[0]),
            "name": str(row[1]),
            "location": str(row[2]),
            "cost": float(row[3]),
            "income": float(row[4]),
            "effect": parse_effect(row[5]),
            "design_intent": str(row[6] or "")
        })
    return dists


def migrate_template_events(wb) -> list[dict]:
    """模板与动力学 → template_events.json"""
    if "模板与动力学" not in wb.sheetnames:
        return []

    templates = []
    for row in rows(wb["模板与动力学"]):
        templates.append({
            "id": str(row[0]),
            "name": str(row[1]),
            "slot": str(row[2]),
            "location": str(row[3]),
            "implicit_effect": str(row[4] or "")
        })
    return templates


def migrate_economy(wb) -> dict:
    """经济与全局参数 → economy.json (economy subset)"""
    if "经济与全局参数" not in wb.sheetnames:
        return {}

    eco_keys = ["初始金钱", "加班收入", "负债压力"]
    eco_map = {
        "初始金钱": "initial_money",
        "加班收入": "overtime_income",
        "负债压力": "debt_stress_per_slot"
    }

    eco = {}
    for row in rows(wb["经济与全局参数"]):
        key = str(row[0])
        if key in eco_keys and isinstance(row[1], (int, float)):
            eco[eco_map[key]] = float(row[1])
    return eco


def migrate_dynamics(wb) -> dict:
    """经济与全局参数 → dynamics.json (dynamics subset)"""
    if "经济与全局参数" not in wb.sheetnames:
        return {}

    dyn_keys = {
        "饱腹消耗/时段": "satiety_drain_per_slot",
        "工作压力增速/时段": "work_stress_per_slot",
        "工作精力消耗/时段": "work_energy_drain",
        "休息降压/时段": "rest_stress_relief",
        "反弹阈值": "rebound_threshold",
        "反弹倍率": "rebound_multiplier",
        "心情耦合速率": "valence_coupling_rate"
    }

    dyn = {}
    for row in rows(wb["经济与全局参数"]):
        key = str(row[0])
        if key in dyn_keys and isinstance(row[1], (int, float)):
            dyn[dyn_keys[key]] = float(row[1])
    return dyn


def migrate_habituation(wb) -> dict:
    """习惯化曲线 → habituation.json"""
    if "习惯化曲线" not in wb.sheetnames:
        return {}

    hab = {}
    for row in rows(wb["习惯化曲线"]):
        if row[0]:
            hab[str(row[0])] = {
                "w_min": float(row[1]),
                "tau": float(row[2]),
                "curve": str(row[3] or "exp")
            }
    return hab


def migrate_needs(wb) -> dict:
    """需求参数 → needs.json"""
    if "需求参数" not in wb.sheetnames:
        return {}

    needs = {}
    for row in rows(wb["需求参数"]):
        if row[0]:
            needs[str(row[0])] = {
                "accumulate": str(row[1] or ""),
                "satisfy_events": str(row[2] or ""),
                "urge_curve": str(row[3] or ""),
                "satisfy_curve": str(row[4] or "")
            }
    return needs


def migrate_persona_modulation(wb) -> dict:
    """人格调节 → persona_modulation.json"""
    if "人格调节" not in wb.sheetnames:
        return {}

    mods = {}
    for row in rows(wb["人格调节"]):
        if row[0]:
            mods[str(row[0])] = {
                "rule": str(row[1] or ""),
                "intent": str(row[2] or "")
            }
    return mods


def save_json(data: any, filename: str) -> None:
    """Save data to JSON file in config/balance/"""
    path = CONFIG_DIR / filename
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"✓ {filename}: {len(data) if isinstance(data, (list, dict)) else 'N/A'} items")


def main() -> None:
    if not XLSX_PATH.exists():
        print(f"❌ Excel 文件不存在: {XLSX_PATH}")
        print("将从代码默认值生成 JSON 配置文件...")
        # Fallback to code defaults - call export_balance_json.py
        import subprocess
        script = PROJECT_ROOT / "scripts" / "export_balance_json.py"
        if script.exists():
            subprocess.run([sys.executable, str(script)])
        return

    from openpyxl import load_workbook

    print(f"正在读取 Excel 配表: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, data_only=True)

    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

    print("\n正在迁移配表到 JSON 文件...")

    # Migrate all sheets
    save_json(migrate_recovery_actions(wb), "recovery_actions.json")
    save_json(migrate_meal_tiers(wb), "meal_tiers.json")
    save_json(migrate_sleep_tiers(wb), "sleep_tiers.json")
    save_json(migrate_custom_activities(wb), "custom_activities.json")
    save_json(migrate_professions(wb), "professions.json")
    save_json(migrate_disturbances(wb), "disturbances.json")
    save_json(migrate_template_events(wb), "template_events.json")
    save_json(migrate_economy(wb), "economy.json")
    save_json(migrate_dynamics(wb), "dynamics.json")
    save_json(migrate_habituation(wb), "habituation.json")
    save_json(migrate_needs(wb), "needs.json")
    save_json(migrate_persona_modulation(wb), "persona_modulation.json")

    print(f"\n✅ 迁移完成！配置文件已保存到: {CONFIG_DIR}")
    print(f"\n原 Excel 文件保留在: {XLSX_PATH}")
    print("建议：备份 Excel 文件后，运行测试验证迁移结果")


if __name__ == "__main__":
    main()

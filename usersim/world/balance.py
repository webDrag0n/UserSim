"""Excel 数值表加载器：balance-sheet/UserSim数值配表.xlsx 为运行时数据源。

加载并覆盖：恢复事件配表 / 扰动事件配表 / 经济与全局参数 / 习惯化曲线 / 需求参数 / 人格调节；
文件或 sheet 缺失时回退代码默认（向后兼容）。进程内缓存 + reload 热更新。
"""

from __future__ import annotations

import re
from pathlib import Path

_cache: dict | None = None

XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "balance-sheet" / "UserSim数值配表.xlsx"

_DIM_EN = {"心情": "valence", "精力": "energy", "饱腹": "satiety", "压力": "stress"}
_PULL_RE = re.compile(r"^(心情|精力|饱腹|压力)→([\d.]+)\(×([\d.]+)\)$")
_EFF_RE = re.compile(r"^(心情|精力|饱腹|压力)([+-][\d.]+)$")


def parse_effect(text: str) -> dict:
    """'心情+0.06 压力-0.04' 或 '饱腹→0.70(×0.75) …' → effect dict。"""
    out: dict = {}
    if not text or text == "—":
        return out
    for part in str(text).split():
        m = _PULL_RE.match(part)
        if m:
            out[_DIM_EN[m.group(1)]] = {"pull": [float(m.group(2)), float(m.group(3))]}
            continue
        m = _EFF_RE.match(part)
        if m:
            out[_DIM_EN[m.group(1)]] = float(m.group(2))
    return out


def _rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(min_row=4, values_only=True) if r and r[0] is not None]


def _load_excel(out: dict) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH, data_only=True)

    if "习惯化曲线" in wb.sheetnames:
        for row in _rows(wb["习惯化曲线"]):
            if row[0]:
                out["habituation"][str(row[0])] = {
                    "w_min": float(row[1]), "tau": float(row[2]), "curve": str(row[3] or "exp"),
                }
        out["source"] = "excel"

    if "需求参数" in wb.sheetnames:
        for row in _rows(wb["需求参数"]):
            if row[0]:
                out["needs"][str(row[0])] = {
                    "accumulate": str(row[1] or ""), "satisfy_events": str(row[2] or ""),
                    "urge_curve": str(row[3] or ""), "satisfy_curve": str(row[4] or ""),
                }
        out["source"] = "excel"

    if "人格调节" in wb.sheetnames:
        for row in _rows(wb["人格调节"]):
            if row[0]:
                out["persona_mod"][str(row[0])] = {"rule": str(row[1] or ""), "intent": str(row[2] or "")}
        out["source"] = "excel"

    if "恢复事件配表" in wb.sheetnames:
        actions: dict[str, dict] = {}
        for row in _rows(wb["恢复事件配表"]):
            aid, action = str(row[0]), str(row[1])
            a = actions.setdefault(aid, {"id": aid, "action": action, "category": str(row[2]),
                                         "design_intent": str(row[11] or ""), "base_effect": parse_effect(row[8]),
                                         "variants": []})
            a["variants"].append({
                "vid": str(row[3]), "location": str(row[4]), "tier": str(row[5]),
                "cost": float(row[6]), "span": int(row[7]),
                "weight": parse_effect(row[9]), "effect": parse_effect(row[10]),
            })
        if actions:
            out["recovery_actions"] = list(actions.values())
            out["source"] = "excel"

    if "扰动事件配表" in wb.sheetnames:
        ds = []
        for row in _rows(wb["扰动事件配表"]):
            ds.append({"id": str(row[0]), "name": str(row[1]), "location": str(row[2]),
                       "cost": float(row[3]), "income": float(row[4]), "effect": parse_effect(row[5]),
                       "design_intent": str(row[6] or "")})
        if ds:
            out["disturbances"] = ds
            out["source"] = "excel"

    if "经济与全局参数" in wb.sheetnames:
        eco = {}
        for row in _rows(wb["经济与全局参数"]):
            if row[0] and isinstance(row[1], (int, float)):
                eco[str(row[0])] = float(row[1])
        if eco:
            out["economy_params"] = eco
            out["source"] = "excel"


def load_overrides(force: bool = False) -> dict:
    """返回覆盖表；结构见各 getter。source ∈ excel|default|default(error)。"""
    global _cache
    if _cache is not None and not force:
        return _cache
    out: dict = {"habituation": {}, "needs": {}, "persona_mod": {}, "source": "default"}
    if XLSX_PATH.exists():
        try:
            _load_excel(out)
        except Exception:
            out["source"] = "default(error)"
    _cache = out
    return out


def reload() -> dict:
    return load_overrides(force=True)

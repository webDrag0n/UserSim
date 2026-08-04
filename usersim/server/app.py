"""FastAPI 后端：运行控制 + 实时推送 + 报告 + 静态托管。

运行模型：每个 run 在后台线程执行（Runner 同步、LLM 调用阻塞），
on_event 回调通过 loop.call_soon_threadsafe 广播给 WebSocket 订阅者。
"""

from __future__ import annotations

import asyncio
import json
import threading
from pathlib import Path

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from usersim.config import PROJECT_ROOT, load_system_config
from usersim.evaluator.report import evaluate_run
from usersim.runner import run_live, run_replay

app = FastAPI(title="UserSim Server")
cfg = load_system_config()
OUT_ROOT = PROJECT_ROOT / str(cfg.run.out_dir)


# ---------------------------------------------------------------
# Run 管理
# ---------------------------------------------------------------


class RunHandle:
    def __init__(self, run_id: str):
        self.run_id = run_id
        self.status = "running"  # running | finished | failed
        self.error: str | None = None
        self.run_dir: Path | None = None
        self.progress = {"slot": 0, "total": 0}
        self.subscribers: list[asyncio.Queue] = []
        self.thread: threading.Thread | None = None


RUNS: dict[str, RunHandle] = {}
MAIN_LOOP: asyncio.AbstractEventLoop | None = None


@app.on_event("startup")
async def _startup() -> None:
    global MAIN_LOOP
    MAIN_LOOP = asyncio.get_running_loop()
    OUT_ROOT.mkdir(parents=True, exist_ok=True)


def _broadcast(handle: RunHandle, event: dict) -> None:
    if MAIN_LOOP is None:
        return

    def _put() -> None:
        for q in list(handle.subscribers):
            try:
                q.put_nowait(event)
            except asyncio.QueueFull:
                pass

    MAIN_LOOP.call_soon_threadsafe(_put)


def _execute(handle: RunHandle, mode: str, seed: int, days: int, quality: str,
             archetype: str | None = None, resume_dir: Path | None = None, extra_days: int = 0,
             harness: str | None = None) -> None:
    def on_event(ev: dict) -> None:
        if ev["type"] == "slot":
            handle.progress["slot"] = ev["data"]["t_logical"] + 1
        _broadcast(handle, ev)

    try:
        if mode == "live":
            run_dir = run_live(seed=seed, days=days, cfg=cfg, out_root=OUT_ROOT, on_event=on_event,
                               archetype=archetype, resume_dir=resume_dir, extra_days=extra_days,
                               run_id=handle.run_id, harness=harness)
        else:
            run_dir = run_replay(seed=seed, days=days, quality=quality, cfg=cfg, out_root=OUT_ROOT,
                                 on_event=on_event, archetype=archetype,
                                 resume_dir=resume_dir, extra_days=extra_days,
                                 run_id=handle.run_id)
        handle.run_dir = run_dir
        report = evaluate_run(run_dir, cfg)
        handle.status = "finished"
        _broadcast(handle, {"type": "done", "data": {"run_id": run_dir.name, "verdict": report["verdict"]}})
    except Exception as e:  # noqa: BLE001
        handle.status = "failed"
        handle.error = str(e)
        _broadcast(handle, {"type": "error", "data": {"error": str(e)}})


class StartRunRequest(BaseModel):
    mode: str = "replay"  # replay | live
    seed: int | None = None
    days: int | None = None
    quality: str = "good"
    archetype: str | None = None
    harness: str | None = None  # 被测 Harness（仅 live）


@app.post("/api/runs")
def start_run(req: StartRunRequest) -> dict:
    seed = req.seed if req.seed is not None else int(cfg.run.seed)
    days = req.days if req.days is not None else int(cfg.run.days)
    # run_id 前置生成：启动即可见、可进入实时观看
    from datetime import datetime, timezone
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
    if req.mode == "live":
        run_id = f"live_{seed}_{ts}"
    else:
        run_id = f"replay_{seed}_{req.quality}_{ts}"
    handle = RunHandle(run_id)
    handle.progress["total"] = days * int(cfg.clock.slots_per_day)
    RUNS[run_id] = handle

    def _wrap() -> None:
        _execute(handle, req.mode, seed, days, req.quality, archetype=req.archetype,
                 harness=req.harness)

    handle.thread = threading.Thread(target=_wrap, daemon=True)
    handle.thread.start()
    return {"started": True, "run_id": run_id, "seed": seed, "days": days, "mode": req.mode}


@app.get("/api/harnesses")
def list_harnesses() -> dict:
    """可选被测 Harness 清单（前端启动表单的下拉数据源）。"""
    from usersim.agents.assistant import DEFAULT_HARNESS, available

    return {"items": available(), "default": DEFAULT_HARNESS}


# ---------------------------------------------------------------
# Bench：多 seed 批量与置信区间
# ---------------------------------------------------------------

BENCH_ROOT = OUT_ROOT / "_bench"
BENCH_JOBS: dict[str, dict] = {}


class StartBenchRequest(BaseModel):
    seeds: str = "1-8"
    days: int = 30
    mode: str = "replay"           # replay | live
    groups: list[str] | None = None
    archetypes: list[str] | None = None
    max_episodes: int | None = None


@app.post("/api/bench")
def start_bench(req: StartBenchRequest) -> dict:
    from usersim.bench import BenchSpec, default_concurrency, estimate_tokens, run_suite
    from usersim.cli import _parse_seeds

    seeds = _parse_seeds(req.seeds)
    groups = req.groups or (["good", "mid", "poor"] if req.mode == "replay" else ["reference"])
    archetypes: list = list(req.archetypes) if req.archetypes else [None]
    spec = BenchSpec(seeds=seeds, days=req.days, mode=req.mode, groups=groups,
                     archetypes=archetypes, concurrency=default_concurrency())
    n_ep = len(spec.episodes())

    if req.mode == "live":
        cap = req.max_episodes
        if cap is None or n_ep > cap:
            return {"started": False, "n_episodes": n_ep,
                    "estimated_tokens": estimate_tokens(spec),
                    "error": "live 批量需显式确认 max_episodes 且不得超过它"}

    from datetime import datetime, timezone
    bench_id = f"bench_{req.mode}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')}"
    job = {"bench_id": bench_id, "status": "running", "done": 0, "total": n_ep, "error": None}
    BENCH_JOBS[bench_id] = job

    def _work() -> None:
        def on_progress(done: int, total: int, last: dict) -> None:
            job["done"] = done
            job["total"] = total

        try:
            run_suite(spec, out_root=BENCH_ROOT, bench_id=bench_id, on_progress=on_progress)
            job["status"] = "finished"
        except Exception as e:  # noqa: BLE001
            job["status"] = "failed"
            job["error"] = str(e)

    threading.Thread(target=_work, daemon=True).start()
    return {"started": True, "bench_id": bench_id, "n_episodes": n_ep}


@app.get("/api/bench")
def list_bench() -> dict:
    """已有批量结果列表（含运行中的任务进度）。"""
    items = []
    if BENCH_ROOT.exists():
        for d in sorted(BENCH_ROOT.iterdir(), reverse=True):
            agg = d / "aggregate.json"
            if not agg.exists():
                continue
            data = json.loads(agg.read_text(encoding="utf-8"))
            items.append({
                "bench_id": d.name,
                "mode": data.get("mode"),
                "days": data.get("days"),
                "n_episodes": data.get("n_episodes"),
                "groups": sorted(data.get("groups", {})),
                "has_guard": (d / "discriminability.json").exists(),
            })
    return {"items": items, "jobs": list(BENCH_JOBS.values())}


@app.get("/api/bench/{bench_id}")
def bench_detail(bench_id: str) -> dict:
    from fastapi import HTTPException

    d = BENCH_ROOT / bench_id
    agg = d / "aggregate.json"
    if not agg.exists():
        job = BENCH_JOBS.get(bench_id)
        if job:
            return {"pending": True, "job": job}
        raise HTTPException(status_code=404, detail=f"未知 bench {bench_id}")
    out: dict = {"aggregate": json.loads(agg.read_text(encoding="utf-8"))}
    guard = d / "discriminability.json"
    if guard.exists():
        out["discriminability"] = json.loads(guard.read_text(encoding="utf-8"))
    eps = d / "episodes.jsonl"
    if eps.exists():
        out["episodes"] = [
            json.loads(l) for l in eps.read_text(encoding="utf-8").splitlines() if l.strip()
        ]
    out["job"] = BENCH_JOBS.get(bench_id)
    return out


class ContinueRunRequest(BaseModel):
    extra_days: int = 10


@app.post("/api/runs/{run_id}/continue")
def continue_run(run_id: str, req: ContinueRunRequest) -> dict:
    """续跑：从 run_state.json 恢复世界，追加 extra_days 天。"""
    d = _find_run_dir(run_id)
    if not (d / "run_state.json").exists():
        from fastapi import HTTPException
        raise HTTPException(400, "该 run 无存档（run_state.json），无法续跑")
    meta = json.loads((d / "meta.json").read_text(encoding="utf-8"))
    total_days = meta["days"] + req.extra_days
    handle = RunHandle(run_id)
    handle.run_dir = d
    handle.progress["total"] = total_days * int(cfg.clock.slots_per_day)
    handle.progress["slot"] = meta["days"] * int(cfg.clock.slots_per_day)
    RUNS[run_id] = handle

    def _wrap() -> None:
        _execute(handle, meta.get("mode", "replay"), meta["seed"], total_days,
                 meta.get("assistant_quality") or "good",
                 resume_dir=d, extra_days=req.extra_days)

    handle.thread = threading.Thread(target=_wrap, daemon=True)
    handle.thread.start()
    return {"continued": True, "run_id": run_id, "extra_days": req.extra_days, "total_days": total_days}


@app.get("/api/balance")
def get_balance() -> dict:
    """配表编辑器数据：可编辑 sheet 的表头与行（含真实行号）。"""
    from openpyxl import load_workbook
    from usersim.world.balance import XLSX_PATH, load_overrides

    editable = ["恢复事件配表", "扰动事件配表", "经济与全局参数", "习惯化曲线", "需求参数", "人格调节", "自定义活动类目"]
    sheets = []
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH, data_only=True)
        for name in editable:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            headers = [c.value for c in ws[3]]
            rows = [[i] + [c.value for c in r] for i, r in enumerate(ws.iter_rows(min_row=4), start=4)
                    if any(c.value is not None for c in r)]
            sheets.append({"name": name, "headers": headers, "rows": rows})
    ov = load_overrides()
    return {"source": ov["source"], "sheets": sheets}


class BalanceCellRequest(BaseModel):
    sheet: str
    row: int
    col: int
    value: str


@app.post("/api/balance/cell")
def set_balance_cell(req: BalanceCellRequest) -> dict:
    """写单元格并热加载（新 run 立即使用新数值）。"""
    from openpyxl import load_workbook
    from usersim.world.balance import XLSX_PATH, reload

    from fastapi import HTTPException
    if not XLSX_PATH.exists():
        raise HTTPException(404, "配表文件不存在")
    wb = load_workbook(XLSX_PATH)
    if req.sheet not in wb.sheetnames:
        raise HTTPException(404, f"sheet 不存在: {req.sheet}")
    ws = wb[req.sheet]
    cell = ws.cell(req.row, req.col)
    try:
        cell.value = float(req.value)  # 数值优先
    except ValueError:
        cell.value = req.value
    wb.save(XLSX_PATH)
    ov = reload()
    return {"ok": True, "source": ov["source"]}


@app.get("/api/runs/{run_id}/insights")
def run_insights(run_id: str) -> dict:
    """深度诊断：故障/拟人性/世界真实性/助手能力的结构化发现。"""
    from usersim.evaluator.insights import compute_insights
    from usersim.evaluator.metrics import load_run

    d = _find_run_dir(run_id)
    slots, turns, meta = load_run(d)
    return compute_insights(slots, turns, meta, cfg.state.targets.to_dict(), float(cfg.state.band),
                            score_cfg=cfg.get("score"))


@app.get("/api/runs/{run_id}/events")
def run_events(run_id: str) -> dict:
    """完整事件记录（模板/扰动/恢复/系列子事件）+ 系列列表，来自世界存档。"""
    d = _find_run_dir(run_id)
    state_file = d / "run_state.json"
    if state_file.exists():
        state = json.loads(state_file.read_text(encoding="utf-8"))
        return {"items": state["world"]["events"], "series": state["world"].get("series", [])}
    # 兼容无存档的旧 run：从 turns 重建恢复事件
    events = []
    f = d / "turns.jsonl"
    if f.exists():
        for l in f.read_text(encoding="utf-8").splitlines():
            t = json.loads(l)
            for r in t.get("tool_results", []):
                if r.get("name") == "add_event_todo" and r.get("ok") and "event" in r.get("payload", {}):
                    events.append(r["payload"]["event"])
    return {"items": events, "series": []}


@app.get("/api/catalog")
def catalog_summary() -> dict:
    """配表摘要（前端参数选择与事件图例用）。"""
    from usersim.world.catalog import MEAL_TIERS, PROFESSIONS, RECOVERY_ACTIONS, SLEEP_TIERS
    return {
        "professions": PROFESSIONS,
        "recovery_actions": [
            {
                "id": a["id"], "action": a["action"], "category": a["category"],
                "variants": [
                    {"vid": v["vid"], "location": v["location"], "tier": v["tier"],
                     "cost": v["cost"], "span": v["span"]}
                    for v in a["variants"]
                ],
            }
            for a in RECOVERY_ACTIONS
        ],
        "meal_tiers": [{"vid": m["vid"], "name": m["name"], "cost": m["cost"]} for m in MEAL_TIERS],
        "sleep_tiers": [{"vid": s["vid"], "name": s["name"], "cost": s["cost"]} for s in SLEEP_TIERS],
    }


@app.get("/api/runs")
def list_runs() -> dict:
    out = []
    for d in sorted(OUT_ROOT.iterdir() if OUT_ROOT.exists() else [], reverse=True):
        meta_file = d / "meta.json"
        if not meta_file.exists():
            continue
        meta = json.loads(meta_file.read_text(encoding="utf-8"))
        report_file = d / "report.json"
        verdict = None
        if report_file.exists():
            verdict = json.loads(report_file.read_text(encoding="utf-8")).get("verdict")
        handle = RUNS.get(d.name)
        persona = meta.get("persona", {})
        item = {
            "run_id": d.name,
            "seed": meta.get("seed"),
            "days": meta.get("days"),
            "mode": meta.get("mode"),
            "assistant_quality": meta.get("assistant_quality"),
            "status": handle.status if handle else "finished",
            "verdict": verdict,
            "persona_name": persona.get("name"),
            "archetype": persona.get("archetype"),
            "income_per_slot": persona.get("income_per_slot"),
            "started_at": meta.get("started_at"),
        }
        if handle is not None and handle.status == "running":
            item["progress"] = dict(handle.progress)
        out.append(item)
    # 进行中的 run（目录/meta 可能尚未创建，列表兜底）
    seen = {r["run_id"] for r in out}
    for rid, h in RUNS.items():
        if h.run_dir is None and rid not in seen:
            out.append({"run_id": rid, "status": h.status, "mode": None, "verdict": None})
    return {"runs": out}


def _find_run_dir(run_id: str) -> Path:
    d = OUT_ROOT / run_id
    if not d.exists():
        from fastapi import HTTPException
        raise HTTPException(404, f"run 不存在: {run_id}")
    return d


class DeleteRunsRequest(BaseModel):
    run_ids: list[str]


@app.post("/api/runs/delete")
def delete_runs(req: DeleteRunsRequest) -> dict:
    """批量删除存档。运行中的 run 拒绝删除；严格限制在 runs 目录内。"""
    import shutil

    from fastapi import HTTPException

    deleted, skipped = [], []
    for rid in req.run_ids:
        # 防路径穿越：run_id 必须是不含分隔符的纯目录名
        if not rid or "/" in rid or "\\" in rid or rid.startswith("."):
            skipped.append({"run_id": rid, "reason": "非法 run_id"})
            continue
        handle = RUNS.get(rid)
        if handle is not None and handle.status == "running":
            skipped.append({"run_id": rid, "reason": "运行中，无法删除"})
            continue
        d = (OUT_ROOT / rid).resolve()
        if not d.is_dir() or OUT_ROOT.resolve() not in d.parents:
            skipped.append({"run_id": rid, "reason": "存档不存在"})
            continue
        RUNS.pop(rid, None)
        shutil.rmtree(d)
        deleted.append(rid)
    return {"deleted": deleted, "skipped": skipped}


@app.get("/api/runs/{run_id}")
def run_detail(run_id: str) -> dict:
    d = _find_run_dir(run_id)
    meta = json.loads((d / "meta.json").read_text(encoding="utf-8"))
    slots_file = d / "slots.jsonl"
    n_slots = sum(1 for _ in slots_file.open()) if slots_file.exists() else 0
    handle = RUNS.get(run_id)
    return {
        "meta": meta,
        "status": handle.status if handle else "finished",
        "error": handle.error if handle else None,
        "n_slots": n_slots,
        "total_slots": meta["days"] * int(cfg.clock.slots_per_day),
    }


@app.get("/api/runs/{run_id}/turns")
def run_turns(run_id: str, offset: int = 0, limit: int = 200) -> dict:
    d = _find_run_dir(run_id)
    f = d / "turns.jsonl"
    lines = f.read_text(encoding="utf-8").splitlines() if f.exists() else []
    items = [json.loads(l) for l in lines[offset:offset + limit]]
    return {"total": len(lines), "offset": offset, "items": items}


@app.get("/api/runs/{run_id}/slots")
def run_slots(run_id: str) -> dict:
    d = _find_run_dir(run_id)
    f = d / "slots.jsonl"
    items = [json.loads(l) for l in f.read_text(encoding="utf-8").splitlines()] if f.exists() else []
    return {"items": items}


@app.get("/api/runs/{run_id}/report")
def run_report(run_id: str) -> dict:
    d = _find_run_dir(run_id)
    report_file = d / "report.json"
    if report_file.exists():
        return json.loads(report_file.read_text(encoding="utf-8"))
    return evaluate_run(d, cfg)


@app.get("/api/config/validation")
def config_validation() -> dict:
    import tomllib
    llm = tomllib.loads((PROJECT_ROOT / "config" / "llm.toml").read_text(encoding="utf-8"))
    providers = {}
    for name, p in llm.get("providers", {}).items():
        key = p.get("api_key", "")
        providers[name] = {
            "model": p.get("model"),
            "key_filled": bool(key) and "在此填入" not in key and len(key) > 10,
        }
    return {
        "system_config_ok": True,
        "providers": providers,
        "roles": {k: v.get("provider") for k, v in llm.get("roles", {}).items()},
    }


# ---------------------------------------------------------------
# WebSocket：先补发积压，再实时推送
# ---------------------------------------------------------------


@app.websocket("/ws/runs/{run_id}")
async def ws_run(ws: WebSocket, run_id: str) -> None:
    await ws.accept()
    d = OUT_ROOT / run_id
    # 补发积压
    if d.exists():
        tf = d / "turns.jsonl"
        if tf.exists():
            for l in tf.read_text(encoding="utf-8").splitlines():
                await ws.send_json({"type": "turn", "data": json.loads(l)})
        sf = d / "slots.jsonl"
        if sf.exists():
            for l in sf.read_text(encoding="utf-8").splitlines():
                await ws.send_json({"type": "slot", "data": json.loads(l)})
    handle = RUNS.get(run_id)
    if handle is None or handle.status != "running":
        await ws.send_json({"type": "done", "data": {"run_id": run_id}})
        await ws.close()
        return
    q: asyncio.Queue = asyncio.Queue(maxsize=2000)
    handle.subscribers.append(q)
    try:
        while True:
            ev = await q.get()
            await ws.send_json(ev)
            if ev["type"] in ("done", "error"):
                break
    except WebSocketDisconnect:
        pass
    finally:
        if q in handle.subscribers:
            handle.subscribers.remove(q)


# ---------------------------------------------------------------
# 静态托管（生产模式：web/dist）
# ---------------------------------------------------------------

DIST = PROJECT_ROOT / str(cfg.frontend.build_dir)
if DIST.exists():
    app.mount("/", StaticFiles(directory=str(DIST), html=True), name="web")


@app.get("/{full_path:path}")
def spa_fallback(full_path: str):
    if DIST.exists():
        candidate = DIST / full_path
        if candidate.is_file():
            return FileResponse(str(candidate))
        return FileResponse(str(DIST / "index.html"))
    return {"hint": "前端未构建：cd web && npm run build，或使用 vite dev 模式"}
